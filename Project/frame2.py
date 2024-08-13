import sqlite3
from tkinter import *
from tkinter import messagebox, ttk
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Initialize the database

# Initialize global variables
product_id = 0
product_id1 = 0
invoice_items = []
update = []

# Database functions
def connect_db():
    return sqlite3.connect('plumbing_store.db')

def execute_query(query, params=()):
    with connect_db() as conn:
        c = conn.cursor()
        c.execute(query, params)
        conn.commit()
def create_table():
    execute_query("""
        CREATE TABLE IF NOT EXISTS products (
            quantity INTEGER NOT NULL,
            price REAL NOT NULL,
            name TEXT NOT NULL,
            id INTEGER PRIMARY KEY AUTOINCREMENT
        )
    """)
create_table()
def fetch_all(query, params=()):
    with connect_db() as conn:
        c = conn.cursor()
        c.execute(query, params)
        return c.fetchall()

# Data validation
def validate_number(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

# CRUD operations
def add_product():
    name = entry_name.get().strip()
    price = entry_price.get().strip()
    quantity = entry_quantity.get().strip()
    
    if not name:
        messagebox.showwarning("تنبيه", "يرجى إدخال اسم المنتج")
        return
    
    if not validate_number(price) or float(price) <= 0:
        messagebox.showwarning("تنبيه", "يرجى إدخال سعر صحيح غير سالب")
        return

    if not quantity.isdigit() or int(quantity) < 0:
        messagebox.showwarning("تنبيه", "يرجى إدخال كمية صحيحة غير سالبة")
        return
    
    price = float(price)
    quantity = int(quantity)
    
    execute_query("INSERT INTO products (name, price, quantity) VALUES (?, ?, ?)", (name, price, quantity))
    messagebox.showinfo("نجاح", "تم إضافة المنتج بنجاح")
    clear_entries()
    show_products()
    show_products1()

def update_product():
    name = entry_name.get().strip()
    price = entry_price.get().strip()
    quantity = entry_quantity.get().strip()

    if not name:
        messagebox.showwarning("تنبيه", "يرجى إدخال اسم المنتج")
        return

    if not validate_number(price) or float(price) <= 0:
        messagebox.showwarning("تنبيه", "يرجى إدخال سعر صحيح غير سالب")
        return

    if not quantity.isdigit() or int(quantity) < 0:
        messagebox.showwarning("تنبيه", "يرجى إدخال كمية صحيحة غير سالبة")
        return
    price = float(price)
    quantity = int(quantity)

    execute_query("UPDATE products SET name=?, price=?, quantity=? WHERE id=?", (name, price, quantity, product_id))
    messagebox.showinfo("نجاح", "تم تحديث المنتج بنجاح")
    clear_entries()
    show_products()
    show_products1()

def delete_product():
    execute_query("DELETE FROM products WHERE id=?", (product_id,))
    messagebox.showinfo("نجاح", "تم حذف المنتج بنجاح")
    clear_entries()
    show_products()
    show_products1()

def show_products():
    for row in tree_inventory.get_children():
        tree_inventory.delete(row)

    rows = fetch_all("SELECT * FROM products")
    
    for row in rows:
        tree_inventory.insert("", "end", values=row)

def on_product_select(event):
    selected_items = tree_inventory.selection()
    if selected_items:
        selected_item = selected_items[0]
        global product_id
        quantity, price, name, product_id = tree_inventory.item(selected_item, 'values')
        entry_name.delete(0, END)
        entry_name.insert(0, name)
        entry_price.delete(0, END)
        entry_price.insert(0, price)
        entry_quantity.delete(0, END)
        entry_quantity.insert(0, quantity)

def search_product():
    search_name = entry_search.get().strip()
    
    for row in tree_inventory.get_children():
        tree_inventory.delete(row)

    rows = fetch_all("SELECT * FROM products WHERE name LIKE ?", ('%' + search_name + '%',))
    for row in rows:
        tree_inventory.insert("", "end", values=row)
    clear_entries()
    clear_entries1()
    
def clear_entries():
    entry_name.delete(0, END)
    entry_price.delete(0, END)
    entry_quantity.delete(0, END)
    entry_search.delete(0, END)

# Sales functions
def clear_entries1():
    entry_name1.delete(0, END)
    entry_discount1.delete(0, END)
    entry_quantity1.delete(0, END)
    entry_search1.delete(0, END)
    entry_person.delete(0, END)
    entry_number.delete(0, END)

def show_products1():
    for row in tree_sales.get_children():
        tree_sales.delete(row)

    with connect_db() as conn:
        c = conn.cursor()
        c.execute("SELECT * FROM products")
        rows = c.fetchall()
    
    for row in rows:
        tree_sales.insert("", "end", values=row)

def sell_product():
    if not entry_quantity1.get() or int(entry_quantity1.get()) <= 0:
        messagebox.showwarning("تنبيه", "يرجى إدخال كمية صحيحة غير سالبة")
        return -1

    discount = 0.0
    if entry_discount1.get():
        if float(entry_discount1.get()) < 0.0 or float(entry_discount1.get()) > 1.0:
            messagebox.showwarning("تنبيه", "يرجى إدخال خصم صحيح غير سالب بين 0 و 1")
            return -1
        else:
            discount = float(entry_discount1.get())

    quantity_to_sell = int(entry_quantity1.get())
    with connect_db() as conn:
        c = conn.cursor()
        c.execute("SELECT name, price, quantity FROM products WHERE id=?", (product_id1,))
        product = c.fetchone()
        if product:
            name, price, current_quantity = product

            if current_quantity >= quantity_to_sell:
                update.append((product_id1, current_quantity))
                new_quantity = current_quantity - quantity_to_sell
                c.execute("UPDATE products SET quantity=? WHERE id=?", (new_quantity, product_id1))
                conn.commit()
                
                discount_price = price * (1 - discount)
                invoice_item = {
                    'اسم الصنف': name,
                    'الكمية': quantity_to_sell,
                    'سعر مبدئي': price,
                    'السعر بعد الخصم': discount_price,
                    'القيمة': discount_price * quantity_to_sell
                }
                invoice_items.append(invoice_item)

                messagebox.showinfo("نجاح", "تم إتمام البيع وتحديث الكمية")
            else:
                messagebox.showwarning("خطأ", "لا يوجد كمية كافية")
        else:
            messagebox.showwarning("خطأ", "المنتج غير موجود")
    
    show_products1()
    show_products()
    clear_entries1()

def on_product_select1(event):
    selected_items = tree_sales.selection()
    if selected_items:
        selected_item = selected_items[0]
        global product_name1, product_price1, product_id1, quantity1  
        quantity1, product_price1, product_name1, product_id1 = tree_sales.item(selected_item, 'values')
        entry_name1.delete(0, END)
        entry_name1.insert(0, product_name1)


def search_product1():
    search_name = entry_search1.get().strip()
    
    for row in tree_sales.get_children():
        tree_sales.delete(row)

    with connect_db() as conn:
        c = conn.cursor()
        c.execute("SELECT * FROM products WHERE name LIKE ?", ('%' + search_name + '%',))
        rows = c.fetchall()
    for row in rows:
        tree_sales.insert("", "end", values=row)
    clear_entries1()

def save_invoice_to_file():
    name = entry_person.get().strip()
    if not name:
        messagebox.showwarning("تنبيه", "يرجى إدخال اسم المشتري")
        return

    phone = entry_number.get().strip()
    if not phone:
        messagebox.showwarning("تنبيه", "يرجى إدخال رقم المشتري")
        return

    op = messagebox.askyesno("هل تريد حفظ الفاتورة ؟", "حفظ")
    if op:
        timestamp = datetime.now().strftime('%Y-%m-%d')  # التاريخ الحالي بتنسيق YYYY-MM-DD
        filename = f"فاتورة.xlsx"

        # إنشاء DataFrame
        levels = [
                    ["الصفا","اسم المحل"],
                    ["01025812907","جمال ابراهيم (المحل)"],
                    ["01090179105","احمد الإمام (السياره)"],
                    [name, 'اسم المشترى'],
                    [phone, 'رقم المشترى'],
                    [timestamp, 'التاريخ'], 
                    ['اسم الصنف', 'الكمية', 'سعر مبدئي', 'السعر بعد الخصم', 'القيمة'] 
                ]
        columns = pd.MultiIndex.from_product(levels, names=["اسم المحل","جمال ابراهيم (المحل)","احمد الإمام (السياره)",'اسم المشترى',"رقم المشترى","التاريخ", 'detail'])

        # تحويل القائمة إلى DataFrame
        df = pd.DataFrame(invoice_items)

        # تعيين MultiIndex للأعمدة
        df.columns = columns[:len(df.columns)]

        # حساب الإجمالي للأعمدة الرقمية فقط
        numeric_columns = df.select_dtypes(include='number').columns
        totals = df[numeric_columns].sum()

        # بناء صف الإجمالي مع وضع 'الإجمالي' في أول عمود فقط
        total_values = ['الإجمالي'] + [totals.get(col, None) for col in numeric_columns]
        # ملء القيم غير الرقمية كـ None
        total_values += [None] * (len(df.columns) - len(total_values))

        # إضافة صف الإجمالي إلى DataFrame
        total_row = pd.Series(total_values, index=df.columns)
        df.loc[len(df)] = total_row

        # حفظ DataFrame إلى ملف Excel بداية من العمود الثاني
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=True, startcol=2)  # يبدأ الحفظ من العمود الثاني (index 1)

        # فتح الملف باستخدام openpyxl وإضافة العنوان في السطر التالي
        workbook = load_workbook(filename)
        sheet = workbook.active

        # الحصول على السطر الأخير للكتابة أسفله
        last_row = sheet.max_row + 3

        # إضافة عنوان المحل في السطر الأخير
        sheet.cell(row=last_row, column=7, value="العنوان: بني عبيد أمام مسجد الشوادي البر الأيسر . الدقهليه")

        # حفظ الملف بعد التعديل
        workbook.save(filename)
        # فتح الملف باستخدام Excel
        os.startfile(filename)
        
        # مسح عناصر الفاتورة بعد الحفظ
        invoice_items.clear()
    else:
        # إذا قام المستخدم بالإلغاء، قم بإرجاع الكميات وعرض رسالة معلومات
        with connect_db() as conn:
            c = conn.cursor()
            for item_id, original_quantity in update:
                c.execute("UPDATE products SET quantity=? WHERE id=?", (original_quantity, item_id))
            conn.commit()
            update.clear()
            messagebox.showinfo("إلغاء", "تم إلغاء الفاتورة واستعادة الكميات الأصلية")

    # Clear entries and refresh the display
    clear_entries1()
    clear_entries()
    show_products1()
    show_products()


# GUI Setup
root = Tk()
root.title("اداره محل الصفا")
root.geometry("1000x600+250+100")
root.resizable(False, False)
# Notebook setup
notebook = ttk.Notebook(root)
notebook.pack(fill=BOTH, expand=True)

# Inventory Tab
tab_inventory = Frame(notebook)
tab_inventory.pack(fill=BOTH, expand=True)
notebook.add(tab_inventory, text='المخزون')

# Sales Tab
tab_sales = Frame(notebook)
tab_sales.pack(fill=BOTH, expand=True)
notebook.add(tab_sales, text='المبيعات')

# Inventory Widgets
frame_inventory = Frame(tab_inventory)
frame_inventory.pack(pady=10)

Label(frame_inventory, text="اسم المنتج").grid(row=0, column=0, padx=10, pady=5)
entry_name = Entry(frame_inventory)
entry_name.grid(row=0, column=1, padx=10, pady=5)

Label(frame_inventory, text="سعر المنتج").grid(row=1, column=0, padx=10, pady=5)
entry_price = Entry(frame_inventory)
entry_price.grid(row=1, column=1, padx=10, pady=5)

Label(frame_inventory, text="الكمية").grid(row=2, column=0, padx=10, pady=5)
entry_quantity = Entry(frame_inventory)
entry_quantity.grid(row=2, column=1, padx=10, pady=5)

Button(frame_inventory, text="إضافة منتج", command=add_product).grid(row=3, column=0, padx=10, pady=10)
Button(frame_inventory, text="تحديث منتج", command=update_product).grid(row=3, column=1, padx=10, pady=10)
Button(frame_inventory, text="حذف منتج", command=delete_product).grid(row=3, column=2, padx=10, pady=10)

Label(frame_inventory, text="بحث عن منتج").grid(row=4, column=0, padx=10, pady=5)
entry_search = Entry(frame_inventory)
entry_search.grid(row=4, column=1, padx=10, pady=5)
Button(frame_inventory, text="بحث", command=search_product).grid(row=4, column=2, padx=10, pady=5)

tree_inventory = ttk.Treeview(tab_inventory, columns=("quantity", "price", "name", "id"), show="headings")
tree_inventory.heading("quantity", text="الكمية")
tree_inventory.heading("price", text="السعر")
tree_inventory.heading("name", text="اسم المنتج")
tree_inventory.heading("id", text="رقم المنتج")
tree_inventory.pack(fill=BOTH, expand=True)
tree_inventory.bind("<<TreeviewSelect>>", on_product_select)

show_products()

# Sales Widgets
frame_sales = Frame(tab_sales)
frame_sales.pack(pady=10)

# Frame for product details
frame_product = Frame(frame_sales)
frame_product.grid(row=0, column=0, padx=10, pady=10, sticky="n")

Label(frame_product, text="اسم المنتج").grid(row=0, column=0, padx=10, pady=5)
entry_name1 = Entry(frame_product)
entry_name1.grid(row=0, column=1, padx=10, pady=5)

Label(frame_product, text="الخصم (%)").grid(row=2, column=0, padx=10, pady=5)
entry_discount1 = Entry(frame_product)
entry_discount1.grid(row=2, column=1, padx=10, pady=5)

Label(frame_product, text="الكمية").grid(row=1, column=0, padx=10, pady=5)
entry_quantity1 = Entry(frame_product)
entry_quantity1.grid(row=1, column=1, padx=10, pady=5)

Button(frame_product, text="بيع منتج", command=sell_product).grid(row=3, column=1, padx=10, pady=10)


Label(frame_product, text="بحث عن منتج").grid(row=4, column=0, padx=10, pady=5)
entry_search1 = Entry(frame_product)
entry_search1.grid(row=4, column=1, padx=10, pady=5)
Button(frame_product, text="بحث", command=search_product1).grid(row=4, column=2, padx=10, pady=5)

tree_sales = ttk.Treeview(tab_sales, columns=("quantity", "price", "name", "id"), show="headings")
tree_sales.heading("quantity", text="الكمية")
tree_sales.heading("price", text="السعر")
tree_sales.heading("name", text="اسم المنتج")
tree_sales.heading("id", text="رقم المنتج")
tree_sales.pack(fill=BOTH, expand=True)
tree_sales.bind("<<TreeviewSelect>>", on_product_select1)

show_products1()

# Frame for customer details
frame_customer = Frame(frame_sales)
frame_customer.grid(row=0, column=1, padx=10, pady=10, sticky="n")

Label(frame_customer, text="اسم المشتري").grid(row=0, column=0, padx=10, pady=5)
entry_person = Entry(frame_customer)
entry_person.grid(row=0, column=1, padx=10, pady=5)

Label(frame_customer, text="رقم المشتري").grid(row=1, column=0, padx=10, pady=5)
entry_number = Entry(frame_customer)
entry_number.grid(row=1, column=1, padx=10, pady=5)

Button(frame_customer, text="حفظ الفاتورة", command=save_invoice_to_file).grid(row=2, column=1, padx=10, pady=10)

root.mainloop()

